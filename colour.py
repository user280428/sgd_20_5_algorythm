import openpyxl
from openpyxl.styles import PatternFill
import os



colours = ['562929', '5d2e2c', '65332e', '6c3830', '743d33',
           '7b4335', '824936', '894f38', '90553a', '965b3b',
           '9d623d', 'a3683e', 'a96f40', 'af7641', 'b47d43',
           'b98544', 'be8c45', 'c39447', 'c79c49', 'cba44b',
           'cfac4d', 'd2b44f', 'd5bd52', 'd7c555', 'd9ce58',
           'dbd65c', 'dcdf60', 'dde865', 'def16a', 'defa70']





# Список наших контуров
def dir_maker():
    lstdr = os.listdir('контуры\\')[2:]
    lstdr = sorted([int(x.split('.')[0].strip('контур')) for x in lstdr])
    lstdr.reverse()

    return lstdr




# Функция
def list_maker(lst=dir_maker()):

    # Общий список, в котором будут другие списки с номерами ячеек, по которым мы будем красить ячейки последнего контура
    main_list = []
    # Пробегаемся по всем эксель файлам
    for i in range(1, len(lst)+1):
        n = len(lst)+1-i
        # Создаем объект таблицы и листа
        wb = openpyxl.load_workbook(f'контуры\\контур{n}.xlsx')
        ws = wb['Sheet1']


        # Пустой лист, куда мы будем записывать координаты ячеек, которые надо закрасить
        cells = []
        # Пробегаемся по рядам, а далее по ячейкам в рядах
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                # Если значение ячейки равно 0, то мы записываем координаты в список
                if cell.internal_value == 0:
                    cells.append(cell.coordinate)
        # Добавляем список в общий список, по которому мы будем красить наши ячейки
        main_list.append(cells)

    return main_list





def colorer_2(colours, main_list=list_maker()):


    for i, l in enumerate(main_list):
        wb = openpyxl.load_workbook(f'контуры\\контур{len(main_list)}.xlsx')
        ws = wb['Sheet1']
        colour = colours[i]
        temp = PatternFill(patternType='solid',
                           fgColor=colour)
        for j in l:
            ws[j].fill = temp
        wb.save(f"контуры\\контур{len(main_list)}.xlsx")

    print('Выполнено')





dir_maker()
list_maker(lst=dir_maker())
colorer_2(colours, main_list=list_maker())